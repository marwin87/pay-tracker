"""Integration tests for GET /export/xlsx — row count and is_deleted exclusion."""

import io

import openpyxl

from tests.conftest import auth, register_and_login

_BILL_A = {
    "name": "Electric",
    "category": "utilities",
    "frequency": "monthly",
    "amount": 100.00,
    "currency": "PLN",
    "due_day": 10,
    "notes": None,
    "is_paused": False,
}

_BILL_B = {
    "name": "Internet",
    "category": "utilities",
    "frequency": "monthly",
    "amount": 60.00,
    "currency": "PLN",
    "due_day": 15,
    "notes": None,
    "is_paused": False,
}


def _data_rows(xlsx_bytes: bytes) -> int:
    """Count total data rows (excluding header) across all sheets in an XLSX workbook."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    return sum(ws.max_row - 1 for ws in wb.worksheets if ws.max_row and ws.max_row > 1)


def test_xlsx_row_count_matches_live_instances(client):
    """XLSX data rows == number of live payment instances for the exported year."""
    tok = register_and_login(client, "xlsx_count@test.com")

    r1 = client.post("/bills", json=_BILL_A, headers=auth(tok))
    assert r1.status_code == 201
    r2 = client.post("/bills", json=_BILL_B, headers=auth(tok))
    assert r2.status_code == 201

    # Trigger instance generation for the current period
    payments = client.get("/bills/payments", headers=auth(tok)).json()
    live_count = len(payments)
    assert live_count >= 2

    r = client.get("/export/xlsx", headers=auth(tok))
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    assert _data_rows(r.content) == live_count


def test_xlsx_excludes_deleted_instances(client):
    """Soft-deleted instances are excluded from the XLSX export (Phase 1 filter check)."""
    tok = register_and_login(client, "xlsx_del@test.com")

    r = client.post("/bills", json=_BILL_A, headers=auth(tok))
    assert r.status_code == 201

    payments = client.get("/bills/payments", headers=auth(tok)).json()
    assert len(payments) == 1
    instance_id = payments[0]["id"]

    # Soft-delete the instance
    r = client.delete(f"/bills/payments/{instance_id}", headers=auth(tok))
    assert r.status_code == 204

    r = client.get("/export/xlsx", headers=auth(tok))
    assert r.status_code == 200
    assert _data_rows(r.content) == 0


def test_xlsx_partial_deletion(client):
    """Deleting one of two instances leaves exactly one row — proves filter is scoped, not blanket."""
    tok = register_and_login(client, "xlsx_partial@test.com")

    r1 = client.post("/bills", json=_BILL_A, headers=auth(tok))
    assert r1.status_code == 201
    r2 = client.post("/bills", json=_BILL_B, headers=auth(tok))
    assert r2.status_code == 201

    payments = client.get("/bills/payments", headers=auth(tok)).json()
    assert len(payments) == 2

    # Soft-delete only one instance
    r = client.delete(f"/bills/payments/{payments[0]['id']}", headers=auth(tok))
    assert r.status_code == 204

    r = client.get("/export/xlsx", headers=auth(tok))
    assert r.status_code == 200
    assert _data_rows(r.content) == 1
